import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from datetime import datetime
import requests
import os
from io import BytesIO

class ReportGenerator:
    def __init__(self):
        self.reports_dir = 'reports'
        os.makedirs(self.reports_dir, exist_ok=True)
    
    def generate_excel_report(self, run_id, run_details, failures):
        """Generate an Excel report with failure details"""
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Failure Report"
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add report header
        ws.merge_cells('A1:F1')
        ws['A1'] = "Test Rigor Failure Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Add run information
        ws['A3'] = "Run ID:"
        ws['B3'] = run_id
        ws['A3'].font = Font(bold=True)
        
        ws['A4'] = "Generated On:"
        ws['B4'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws['A4'].font = Font(bold=True)
        
        ws['A5'] = "Total Failed Tests:"
        ws['B5'] = len(failures)
        ws['A5'].font = Font(bold=True)
        ws['B5'].font = Font(color="FF0000", bold=True)
        
        if run_details:
            ws['D3'] = "Run Name:"
            ws['E3'] = run_details.get('name', 'N/A')
            ws['D3'].font = Font(bold=True)
            
            ws['D4'] = "Status:"
            ws['E4'] = run_details.get('status', 'N/A')
            ws['D4'].font = Font(bold=True)
        
        # Add table headers
        headers = ['Test Case ID', 'Test Name', 'Status', 'Screenshot Number', 'Failed Command', 'Error Message', 'Screenshot']
        header_row = 7
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add failure data
        current_row = header_row + 1
        
        for failure in failures:
            ws.cell(row=current_row, column=1).value = failure.get('test_case_id', 'N/A')
            ws.cell(row=current_row, column=2).value = failure.get('test_name', 'N/A')
            ws.cell(row=current_row, column=3).value = failure.get('status', 'N/A')
            ws.cell(row=current_row, column=4).value = failure.get('screenshot_number', 'N/A')
            ws.cell(row=current_row, column=5).value = failure.get('failed_command', 'N/A')
            ws.cell(row=current_row, column=6).value = failure.get('error_message', 'N/A')
            
            # Add screenshot link or download screenshot
            screenshot_url = failure.get('screenshot_url')
            if screenshot_url:
                ws.cell(row=current_row, column=7).value = screenshot_url
                ws.cell(row=current_row, column=7).hyperlink = screenshot_url
                ws.cell(row=current_row, column=7).font = Font(color="0000FF", underline="single")
            else:
                ws.cell(row=current_row, column=7).value = 'N/A'
            
            # Apply borders to all cells in the row
            for col in range(1, 8):
                ws.cell(row=current_row, column=col).border = border
                ws.cell(row=current_row, column=col).alignment = Alignment(vertical='top', wrap_text=True)
            
            current_row += 1
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 50
        ws.column_dimensions['G'].width = 20
        
        # Set row height for data rows to accommodate wrapped text
        for row in range(header_row + 1, current_row):
            ws.row_dimensions[row].height = 60
        
        # Save the workbook
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"TestRigor_FailureReport_{run_id}_{timestamp}.xlsx"
        filepath = os.path.join(self.reports_dir, filename)
        
        wb.save(filepath)
        
        return filename
    
    def download_screenshot(self, url, test_id):
        """Download screenshot from URL"""
        try:
            response = requests.get(url)
            if response.status_code == 200:
                filename = f"screenshot_{test_id}.png"
                filepath = os.path.join(self.reports_dir, filename)
                with open(filepath, 'wb') as f:
                    f.write(response.content)
                return filepath
            return None
        except Exception as e:
            print(f"Error downloading screenshot: {str(e)}")
            return None
