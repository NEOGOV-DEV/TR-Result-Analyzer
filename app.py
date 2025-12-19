from flask import Flask, request, jsonify, render_template, send_file
from flask_cors import CORS
import requests
import json
from datetime import datetime
import os
from concurrent.futures import ThreadPoolExecutor, as_completed
from report_generator import ReportGenerator

app = Flask(__name__)
CORS(app)

# Load configuration
with open('config.json', 'r') as f:
    config = json.load(f)

SUITES = config.get('suites', {})

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/api/generate-report', methods=['POST'])
def generate_report():
    try:
        data = request.json
        suite = data.get('suite')
        run_id = data.get('run_id')
        
        if not suite:
            return jsonify({'error': 'Suite is required'}), 400
        
        if not run_id:
            return jsonify({'error': 'Run ID is required'}), 400
        
        # Get suite configuration
        suite_config = SUITES.get(suite)
        if not suite_config:
            return jsonify({'error': f'Invalid suite: {suite}'}), 400
        
        api_key = suite_config.get('auth_token')
        base_url = suite_config.get('base_url')
        
        if not api_key or not base_url:
            return jsonify({'error': f'Suite configuration incomplete for: {suite}'}), 400
        
        # Fetch test run details from Test Rigor API
        run_details, error_msg, status_code = fetch_run_details(run_id, api_key, base_url)
        
        if not run_details:
            return jsonify({'error': error_msg}), status_code
        
        # Fetch test cases in the run
        test_cases = fetch_test_cases(run_id, api_key, base_url)
        
        # Process failures
        failures = process_failures(run_id, test_cases, api_key, base_url)
        
        # Generate report
        report_generator = ReportGenerator()
        report_path = report_generator.generate_excel_report(run_id, run_details, failures, api_key, suite_name=suite)
        
        # Prepare report data for UI display
        report_data = {
            'suite_name': suite,
            'run_id': run_id,
            'run_name': run_details.get('customName', 'N/A') if run_details else 'N/A',
            'status': run_details.get('status', 'N/A') if run_details else 'N/A',
            'generated_on': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'failures': []
        }
        
        # Format failures for display with "No command found" processing
        for failure in failures:
            failed_command = failure.get('failed_command', 'N/A')
            
            # Apply "No command found" logic for empty commands
            if failed_command and failed_command != 'N/A':
                lines = failed_command.split('\n')
                processed_lines = []
                
                for line in lines:
                    # Check if line is just "Step X:" without any command
                    if line.strip() and ':' in line:
                        parts = line.split(':', 1)
                        if len(parts) == 2 and parts[0].strip().startswith('Step'):
                            command_part = parts[1].strip()
                            if not command_part:
                                processed_lines.append(f"{parts[0].strip()}: No command found")
                            else:
                                processed_lines.append(line)
                        else:
                            processed_lines.append(line)
                    elif line.strip():
                        processed_lines.append(line)
                
                processed_command = '\n'.join(processed_lines).strip()
                
                # If after processing we only have whitespace, use "No command found"
                if not processed_command or processed_command.isspace():
                    failed_command = 'No command found'
                else:
                    failed_command = processed_command
            
            report_data['failures'].append({
                'test_case_id': failure.get('test_case_id', 'N/A'),
                'test_case_uuid': failure.get('test_case_uuid', ''),
                'test_name': failure.get('test_name', 'N/A'),
                'status': failure.get('status', 'N/A'),
                'screenshot_number': failure.get('screenshot_number', 'N/A'),
                'failed_command': failed_command,
                'error_message': failure.get('error_message', 'N/A'),
                'screenshot_urls': failure.get('screenshot_urls', [])
            })
        
        # Add suite_name for suite_id mapping in frontend
        report_data['suite_name'] = suite
        
        return jsonify({
            'success': True,
            'message': 'Report generated successfully',
            'report_path': report_path,
            'total_tests': len(test_cases),
            'failed_tests': len(failures),
            'report_data': report_data
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/download-report/<filename>')
def download_report(filename):
    try:
        report_path = os.path.join('reports', filename)
        if os.path.exists(report_path):
            return send_file(report_path, as_attachment=True)
        else:
            return jsonify({'error': 'Report not found'}), 404
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/compare-runs', methods=['POST'])
def compare_runs():
    """Get all test cases with statuses for comparison (not just failures)"""
    try:
        data = request.json
        suite = data.get('suite')
        run_id = data.get('run_id')
        
        if not suite:
            return jsonify({'error': 'Suite is required'}), 400
        
        if not run_id:
            return jsonify({'error': 'Run ID is required'}), 400
        
        # Get suite configuration
        suite_config = SUITES.get(suite)
        if not suite_config:
            return jsonify({'error': f'Invalid suite: {suite}'}), 400
        
        api_key = suite_config.get('auth_token')
        base_url = suite_config.get('base_url')
        
        if not api_key or not base_url:
            return jsonify({'error': f'Suite configuration incomplete for: {suite}'}), 400
        
        # Fetch test run details
        run_details, error_msg, status_code = fetch_run_details(run_id, api_key, base_url)
        
        if not run_details:
            return jsonify({'error': error_msg}), status_code
        
        # Fetch all test cases in the run
        test_cases = fetch_test_cases(run_id, api_key, base_url)
        
        # Extract all test cases with their statuses
        all_tests = []
        for test in test_cases:
            all_tests.append({
                'test_case_id': test.get('referenceName', test.get('name', 'N/A')),
                'test_name': test.get('name', 'N/A'),
                'status': test.get('status', 'N/A')
            })
        
        return jsonify({
            'success': True,
            'run_name': run_details.get('customName', run_details.get('name', 'N/A')),
            'run_date': run_details.get('startTime', run_details.get('createdAt', 'N/A')),
            'total_tests': len(test_cases),
            'all_tests': all_tests
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/download-summary-report', methods=['POST'])
def download_summary_report():
    """Generate Excel report with Overall Summary and Detailed Summary"""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from io import BytesIO
        
        data = request.json
        summary = data.get('summary', {})
        detailed_tests = data.get('detailed_tests', [])
        
        # Create workbook
        wb = Workbook()
        
        # Sheet 1: Overall Summary
        ws1 = wb.active
        ws1.title = "Overall Summary"
        
        # Header styling
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Title
        ws1['A1'] = 'Overall Summary Report'
        ws1['A1'].font = Font(bold=True, size=14)
        ws1.merge_cells('A1:H1')
        
        ws1['A2'] = f"Suite: {summary.get('suite', 'N/A')}"
        ws1['A2'].font = Font(bold=True)
        ws1.merge_cells('A2:H2')
        
        ws1['A3'] = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        ws1.merge_cells('A3:H3')
        
        # Helper function to format date/time
        def format_date_time(date_str):
            if not date_str or date_str == 'N/A':
                return 'N/A'
            try:
                # Parse ISO format datetime
                if 'T' in date_str:
                    dt = datetime.fromisoformat(date_str.replace('Z', '+00:00'))
                else:
                    dt = datetime.strptime(date_str, '%Y-%m-%d %H:%M:%S')
                # Format as "Dec 16, 2025, 01:47 PM"
                return dt.strftime('%b %d, %Y, %I:%M %p')
            except:
                return date_str
        
        # Headers for summary table
        headers = ['Run Name', 'Run ID', 'Run Date/Time', 'Total Tests', 'Passed Tests', 'Failed Tests', 'Valid Fail', 'RTR']
        ws1.append([])  # Empty row
        header_row = ws1.max_row + 1
        
        for col, header in enumerate(headers, start=1):
            cell = ws1.cell(row=header_row, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Add run data
        for run in summary.get('runs', []):
            passed_tests = run.get('total_tests', 0) - run.get('failed_tests', 0)
            ws1.append([
                run.get('run_name', 'N/A'),
                run.get('run_id', 'N/A'),
                format_date_time(run.get('run_date', 'N/A')),
                run.get('total_tests', 0),
                passed_tests,
                run.get('failed_tests', 0),
                run.get('valid_fail', 0),
                run.get('rtr', '0%')
            ])
            
            # Apply border to data cells
            for col in range(1, 9):
                ws1.cell(row=ws1.max_row, column=col).border = border
                ws1.cell(row=ws1.max_row, column=col).alignment = Alignment(horizontal='center')
        
        # Adjust column widths
        ws1.column_dimensions['A'].width = 25
        ws1.column_dimensions['B'].width = 20
        ws1.column_dimensions['C'].width = 25
        ws1.column_dimensions['D'].width = 12
        ws1.column_dimensions['E'].width = 12
        ws1.column_dimensions['F'].width = 12
        ws1.column_dimensions['G'].width = 12
        ws1.column_dimensions['H'].width = 10
        
        # Sheet 2: Detailed Summary
        ws2 = wb.create_sheet(title="Detailed Summary")
        
        # Determine number of runs
        num_runs = len(summary.get('runs', []))
        
        # Headers for detailed table
        detail_headers = ['Test Case ID', 'Test Name']
        for i, run in enumerate(summary.get('runs', []), start=1):
            detail_headers.append(f"{run.get('run_name', f'Run {i}')} Status")
        
        # Add headers
        for col, header in enumerate(detail_headers, start=1):
            cell = ws2.cell(row=1, column=col, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
        
        # Add test data
        for test in detailed_tests:
            row_data = [
                test.get('test_case_id', 'N/A'),
                test.get('test_name', 'N/A')
            ]
            
            # Add all run statuses (up to 7 runs)
            for i in range(1, 8):
                run_status_key = f'run{i}_status'
                if run_status_key in test:
                    row_data.append(test.get(run_status_key, 'N/A'))
            
            ws2.append(row_data)
            
            # Apply border and alignment
            for col in range(1, len(row_data) + 1):
                cell = ws2.cell(row=ws2.max_row, column=col)
                cell.border = border
                if col > 2:  # Status columns
                    cell.alignment = Alignment(horizontal='center')
                    # Color code status
                    status = cell.value.lower() if cell.value else ''
                    if 'passed' in status:
                        cell.fill = PatternFill(start_color='C6EFCE', end_color='C6EFCE', fill_type='solid')
                    elif 'failed' in status or 'error' in status:
                        cell.fill = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
        
        # Adjust column widths for detailed sheet
        ws2.column_dimensions['A'].width = 15
        ws2.column_dimensions['B'].width = 50
        for i in range(3, len(detail_headers) + 1):
            ws2.column_dimensions[chr(64 + i)].width = 20
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return as file download
        filename = f"Summary_Report_{summary.get('suite', 'Unknown')}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/load-failed-tests', methods=['POST'])
def load_failed_tests():
    """Optimized endpoint to load failed tests for Rerun tab (without detailed processing)"""
    try:
        data = request.json
        suite = data.get('suite')
        run_id = data.get('run_id')
        
        if not suite:
            return jsonify({'error': 'Suite is required'}), 400
        
        if not run_id:
            return jsonify({'error': 'Run ID is required'}), 400
        
        # Get suite configuration
        suite_config = SUITES.get(suite)
        if not suite_config:
            return jsonify({'error': f'Invalid suite: {suite}'}), 400
        
        api_key = suite_config.get('auth_token')
        base_url = suite_config.get('base_url')
        
        if not api_key or not base_url:
            return jsonify({'error': f'Suite configuration incomplete for: {suite}'}), 400
        
        # Fetch test run details
        run_details, error_msg, status_code = fetch_run_details(run_id, api_key, base_url)
        
        if not run_details:
            return jsonify({'error': error_msg}), status_code
        
        # Fetch test cases (this is fast - just one API call with pagination)
        test_cases = fetch_test_cases(run_id, api_key, base_url)
        
        # Process failures with minimal processing (no execution details, screenshots, or commands)
        failures = process_failures_minimal(test_cases)
        
        # Prepare report data for UI
        report_data = {
            'suite_name': suite,
            'run_id': run_id,
            'run_name': run_details.get('customName', 'N/A') if run_details else 'N/A',
            'status': run_details.get('status', 'N/A') if run_details else 'N/A',
            'generated_on': datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            'failures': failures
        }
        
        return jsonify({
            'success': True,
            'message': 'Failed tests loaded successfully',
            'total_tests': len(test_cases),
            'failed_tests': len(failures),
            'report_data': report_data
        })
    
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/rerun-failed', methods=['POST'])
def rerun_failed_tests():
    """Trigger rerun of failed test cases from a specific run"""
    try:
        data = request.json
        suite = data.get('suite')
        run_id = data.get('run_id')
        custom_name = data.get('custom_name', f'Rerun of failed tests - {run_id}')
        
        if not suite:
            return jsonify({'error': 'Suite is required'}), 400
        
        if not run_id:
            return jsonify({'error': 'Run ID is required'}), 400
        
        # Get suite configuration
        suite_config = SUITES.get(suite)
        if not suite_config:
            return jsonify({'error': f'Invalid suite: {suite}'}), 400
        
        api_key = suite_config.get('auth_token')
        base_url = suite_config.get('base_url')
        
        if not api_key or not base_url:
            return jsonify({'error': f'Suite configuration incomplete for: {suite}'}), 400
        
        # Fetch run details to get the application URL
        run_details, error_msg, status_code = fetch_run_details(run_id, api_key, base_url)
        
        if not run_details:
            return jsonify({'error': f'Failed to fetch run details: {error_msg}'}), status_code
        
        # Get the application URL from run details
        app_url = run_details.get('url', run_details.get('applicationUrl', ''))
        
        if not app_url:
            return jsonify({'error': 'Application URL not found in run details. Cannot trigger rerun.'}), 400
        
        # Fetch test cases from the run
        test_cases = fetch_test_cases(run_id, api_key, base_url)
        
        if not test_cases:
            return jsonify({'error': 'No test cases found for this run'}), 400
        
        # Extract UUIDs of failed test cases
        failed_uuids = []
        for test in test_cases:
            status = test.get('status', '')
            if status in ['Failed', 'failed', 'error', 'Error']:
                test_case_uuid = test.get('testCaseUuid')
                if test_case_uuid:
                    failed_uuids.append(test_case_uuid)
        
        if not failed_uuids:
            return jsonify({'error': 'No failed test cases found in this run'}), 400
        
        # Trigger rerun via TestRigor API
        # Note: Rerun endpoint uses api.testrigor.com instead of api2.testrigor.com
        headers = {
            'auth-token': api_key,
            'Content-Type': 'application/json'
        }
        
        rerun_data = {
            'testCaseUuids': failed_uuids,
            'customName': custom_name,
            'url': app_url
        }
        
        # Extract app ID and construct rerun URL with correct base
        app_id = base_url.split('/apps/')[-1]
        rerun_base_url = f'https://api.testrigor.com/api/v1/apps/{app_id}'
        url = f'{rerun_base_url}/retest'
        print(f"\n=== TRIGGERING RERUN ===")
        print(f"URL: {url}")
        print(f"Failed test cases to rerun: {len(failed_uuids)}")
        print(f"Request data: {json.dumps(rerun_data, indent=2)}")
        
        # Send as JSON
        response = requests.post(url, headers=headers, json=rerun_data)
        
        print(f"Status Code: {response.status_code}")
        print(f"Response: {response.text}")
        
        if response.status_code in [200, 201]:
            response_json = response.json()
            return jsonify({
                'success': True,
                'message': f'Successfully triggered rerun of {len(failed_uuids)} failed test(s)',
                'failed_tests_count': len(failed_uuids),
                'task_id': response_json.get('taskId'),
                'queue_id': response_json.get('queueId'),
                'result': response_json.get('result')
            })
        else:
            return jsonify({
                'error': f'Failed to trigger rerun: {response.text}'
            }), response.status_code
    
    except Exception as e:
        print(f"Exception in rerun_failed_tests: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/rerun-selected', methods=['POST'])
def rerun_selected_tests():
    """Trigger rerun of specific selected test cases"""
    try:
        data = request.json
        suite = data.get('suite')
        run_id = data.get('run_id')
        test_case_uuids = data.get('test_case_uuids', [])
        custom_name = data.get('custom_name', f'Rerun selected tests - {run_id}')
        
        if not suite:
            return jsonify({'error': 'Suite is required'}), 400
        
        if not run_id:
            return jsonify({'error': 'Run ID is required'}), 400
        
        if not test_case_uuids or len(test_case_uuids) == 0:
            return jsonify({'error': 'No test case UUIDs provided'}), 400
        
        # Get suite configuration
        suite_config = SUITES.get(suite)
        if not suite_config:
            return jsonify({'error': f'Invalid suite: {suite}'}), 400
        
        api_key = suite_config.get('auth_token')
        base_url = suite_config.get('base_url')
        
        if not api_key or not base_url:
            return jsonify({'error': f'Suite configuration incomplete for: {suite}'}), 400
        
        # Fetch run details to get the application URL
        run_details, error_msg, status_code = fetch_run_details(run_id, api_key, base_url)
        
        if not run_details:
            return jsonify({'error': f'Failed to fetch run details: {error_msg}'}), status_code
        
        # Get the application URL from run details
        app_url = run_details.get('url', run_details.get('applicationUrl', ''))
        
        if not app_url:
            return jsonify({'error': 'Application URL not found in run details. Cannot trigger rerun.'}), 400
        
        # Trigger rerun via TestRigor API
        headers = {
            'auth-token': api_key,
            'Content-Type': 'application/json'
        }
        
        rerun_data = {
            'testCaseUuids': test_case_uuids,
            'customName': custom_name,
            'url': app_url
        }
        
        # Extract app ID and construct rerun URL
        app_id = base_url.split('/apps/')[-1]
        rerun_base_url = f'https://api.testrigor.com/api/v1/apps/{app_id}'
        url = f'{rerun_base_url}/retest'
        
        print(f"\n=== TRIGGERING SELECTED TESTS RERUN ===")
        print(f"URL: {url}")
        print(f"Selected test cases to rerun: {len(test_case_uuids)}")
        print(f"Request data: {json.dumps(rerun_data, indent=2)}")
        
        response = requests.post(url, headers=headers, json=rerun_data)
        
        print(f"Status Code: {response.status_code}")
        print(f"Response: {response.text}")
        
        if response.status_code in [200, 201]:
            response_json = response.json()
            return jsonify({
                'success': True,
                'message': f'Successfully triggered rerun of {len(test_case_uuids)} selected test(s)',
                'failed_tests_count': len(test_case_uuids),
                'task_id': response_json.get('taskId'),
                'queue_id': response_json.get('queueId'),
                'result': response_json.get('result')
            })
        else:
            return jsonify({
                'error': f'Failed to trigger rerun: {response.text}'
            }), response.status_code
    
    except Exception as e:
        print(f"Exception in rerun_selected_tests: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

def fetch_run_details(run_id, api_key, base_url):
    """Fetch test run details from Test Rigor API"""
    try:
        headers = {
            'auth-token': api_key
        }
        
        url = f'{base_url}/runs/{run_id}'
        print(f"\n=== FETCHING RUN DETAILS ===")
        print(f"URL: {url}")
        
        # Extract expected application ID from base_url
        expected_app_id = base_url.split('/apps/')[-1]
        print(f"Expected Application ID: {expected_app_id}")
        
        response = requests.get(url, headers=headers)
        
        print(f"Status Code: {response.status_code}")
        
        if response.status_code == 200:
            response_json = response.json()
            print(f"Response Keys: {list(response_json.keys())}")
            data = response_json.get('data', response_json)
            
            # Validate that the run belongs to the selected suite/application
            actual_app_id = data.get('applicationId')
            print(f"Actual Application ID from response: {actual_app_id}")
            
            if actual_app_id and actual_app_id != expected_app_id:
                print(f"Application ID mismatch! Expected: {expected_app_id}, Got: {actual_app_id}")
                return None, 'Suite selected and provided Run ID does not match', 400
            
            print(f"\n=== RUN DETAILS FIELDS ===")
            print(f"Available fields in run details: {list(data.keys())}")
            print(f"Run Details: {json.dumps(data, indent=2)[:1000]}")
            return data, None, 200
        elif response.status_code == 404:
            print(f"Error: {response.status_code} - {response.text}")
            return None, 'Suite selected and provided Run ID does not match', 400
        elif response.status_code in [401, 403]:
            print(f"Error: {response.status_code} - {response.text}")
            return None, 'Authentication failed. Please check API credentials in config.json', 401
        else:
            print(f"Error: {response.status_code} - {response.text}")
            return None, f'Failed to fetch run details. Status code: {response.status_code}', 400
    
    except Exception as e:
        print(f"Exception in fetch_run_details: {str(e)}")
        import traceback
        traceback.print_exc()
        return None, f'Error connecting to Test Rigor API: {str(e)}', 500

def fetch_test_cases(run_id, api_key, base_url):
    """Fetch all test cases for a given run with pagination"""
    try:
        headers = {
            'auth-token': api_key
        }
        
        print(f"\n=== FETCHING TEST CASES ===")
        all_test_cases = []
        page = 0
        page_size = 100
        
        while True:
            url = f'{base_url}/runs/{run_id}/testcases?page={page}&size={page_size}'
            print(f"Fetching page {page}: {url}")
            
            response = requests.get(url, headers=headers)
            print(f"Status Code: {response.status_code}")
            
            if response.status_code == 200:
                response_json = response.json()
                data = response_json.get('data', {})
                test_cases = data.get('content', [])
                
                print(f"Test Cases on page {page}: {len(test_cases)}")
                
                if not test_cases:
                    break
                
                all_test_cases.extend(test_cases)
                
                # Check if there are more pages
                total_elements = data.get('totalElements', 0)
                print(f"Total elements: {total_elements}, Fetched so far: {len(all_test_cases)}")
                
                if len(all_test_cases) >= total_elements:
                    break
                
                page += 1
            else:
                print(f"Error: {response.status_code} - {response.text}")
                break
        
        print(f"Total Test Cases Fetched: {len(all_test_cases)}")
        return all_test_cases
    
    except Exception as e:
        print(f"Exception in fetch_test_cases: {str(e)}")
        import traceback
        traceback.print_exc()
        return []

def process_single_failure(run_id, test, api_key, base_url):
    """Process a single failed test case - used for parallel processing"""
    import re
    
    try:
        # Extract Test Case ID from name (C followed by numbers)
        test_name = test.get('name', 'Unknown Test')
        match = re.search(r'C\d+', test_name)
        test_case_id = match.group(0) if match else 'N/A'
        
        # Get execution details
        test_case_uuid = test.get('testCaseUuid')
        executions = test.get('executions', [])
        status = test.get('status', '')
        
        failure_step = 'N/A'
        error_message = 'No error details available'
        screenshot_url = None
        failed_command = 'N/A'
        screenshot_urls = []
        
        if not test_case_uuid:
            print(f"Warning: No test_case_uuid for {test_case_id}")
            error_message = 'No test case UUID available'
        elif not executions or len(executions) == 0:
            print(f"Warning: No executions for {test_case_id}")
            error_message = 'No execution data available'
        else:
            execution_uuid = executions[0].get('uuid')
            
            if not execution_uuid:
                print(f"Warning: No execution UUID for {test_case_id}")
                error_message = 'No execution UUID available'
            else:
                # Fetch detailed execution info to get failed step
                print(f"Fetching execution details for {test_case_id}...")
                execution_details = fetch_execution_details(run_id, test_case_uuid, execution_uuid, api_key, base_url)
                
                if execution_details:
                    # execution_details is either a dict with 'steps' key or directly a list of steps
                    steps = execution_details if isinstance(execution_details, list) else execution_details.get('steps', [])
                    
                    for step in steps:
                        step_status = step.get('status')
                        
                        if step_status in ['Failed', 'failed', 'Error', 'error']:
                            failure_step = step.get('step', 'N/A')
                            
                            # Get the error message (full stepDescription)
                            error_message = step.get('stepDescription', 'No error details available')
                            
                            # Collect commands from failure step and 3 steps before
                            failed_commands = []
                            start_step = max(1, failure_step - 3)
                            
                            for i, s in enumerate(steps):
                                s_num = s.get('step', i + 1)
                                if start_step <= s_num <= failure_step:
                                    cmd = (
                                        s.get('command') or 
                                        s.get('action') or 
                                        s.get('instruction') or 
                                        s.get('testCommand') or 
                                        s.get('scriptLine') or
                                        s.get('commandText') or
                                        s.get('stepDescription', 'N/A')
                                    )
                                    failed_commands.append(f"Step {s_num}: {cmd}")
                            
                            # Join all commands with line breaks
                            failed_command = "\n".join(failed_commands) if failed_commands else 'N/A'
                            
                            # Get screenshot URL from step data - check multiple possible locations
                            screenshot_url = None
                            
                            # Check direct fields
                            for field in ['screenshot', 'screenshotUrl', 'screenshotLink', 'image', 'imageUrl', 'screenshotPath']:
                                if step.get(field):
                                    screenshot_url = step.get(field)
                                    break
                            
                            # Check nested in errors array
                            if not screenshot_url:
                                errors = step.get('errors', [])
                                for error in errors:
                                    if isinstance(error, dict):
                                        for field in ['screenshot', 'screenshotUrl', 'image']:
                                            if error.get(field):
                                                screenshot_url = error.get(field)
                                                break
                                    if screenshot_url:
                                        break
                            
                            # TestRigor stores screenshots in S3 with a specific naming pattern
                            # Construct the S3 URLs for current step and 3 steps before
                            screenshot_urls = []
                            if not screenshot_url and execution_uuid and failure_step:
                                # Generate URLs for current step and up to 3 steps before
                                start_step = max(1, failure_step - 3)
                                urls_to_validate = []
                                for step_num in range(start_step, failure_step + 1):
                                    url = f"https://files-to-test.s3.amazonaws.com/web-app_{execution_uuid}_{step_num}.png"
                                    urls_to_validate.append(url)
                                
                                # Validate all URLs concurrently for better performance
                                validated_results = validate_screenshots_batch(urls_to_validate)
                                
                                # Filter out None values (unavailable screenshots)
                                screenshot_urls = [url for url in validated_results if url is not None]
                                screenshot_url = screenshot_urls[0] if screenshot_urls else None
                            
                            break
        
        failure_info = {
            'test_case_id': test_case_id,
            'test_case_uuid': test_case_uuid,
            'test_name': test_name,
            'status': status,
            'screenshot_number': failure_step,
            'failed_command': failed_command,
            'error_message': error_message,
            'screenshot_url': screenshot_url,
            'screenshot_urls': screenshot_urls
        }
        
        print(f"Processed failure: {test_case_id} - {test_name}")
        return failure_info
        
    except Exception as e:
        # If any error occurs, return minimal info
        print(f"Exception in process_single_failure for {test.get('name', 'Unknown')}: {str(e)}")
        import traceback
        traceback.print_exc()
        
        import re
        test_name = test.get('name', 'Unknown Test')
        match = re.search(r'C\d+', test_name)
        test_case_id = match.group(0) if match else 'N/A'
        
        return {
            'test_case_id': test_case_id,
            'test_case_uuid': test.get('testCaseUuid'),
            'test_name': test_name,
            'status': test.get('status', ''),
            'screenshot_number': 'N/A',
            'failed_command': 'N/A',
            'error_message': f'Error processing: {str(e)}',
            'screenshot_url': None,
            'screenshot_urls': []
        }

def process_failures(run_id, test_cases, api_key, base_url):
    """Process failed test cases and extract failure details with parallel processing"""
    failures = []
    
    print(f"\n=== PROCESSING FAILURES (Parallel) ===")
    print(f"Total test cases: {len(test_cases)}")
    
    # Filter out only failed test cases
    failed_tests = [test for test in test_cases if test.get('status', '') in ['Failed', 'failed', 'error', 'Error']]
    
    print(f"Failed test cases to process: {len(failed_tests)}")
    
    if not failed_tests:
        print("No failures to process")
        return []
    
    # Process failures in parallel using ThreadPoolExecutor
    # Use max_workers=3 to avoid API rate limits (reduced from 5)
    with ThreadPoolExecutor(max_workers=3) as executor:
        # Submit all failure processing tasks
        future_to_test = {
            executor.submit(process_single_failure, run_id, test, api_key, base_url): test 
            for test in failed_tests
        }
        
        # Collect results as they complete
        for future in as_completed(future_to_test):
            test = future_to_test[future]
            try:
                failure_info = future.result(timeout=30)  # Add 30 second timeout
                failures.append(failure_info)
            except Exception as e:
                # If processing fails, create a minimal failure record
                test_name = test.get('name', 'Unknown')
                print(f"Error processing failure for {test_name}: {str(e)}")
                import traceback
                traceback.print_exc()
                
                # Add minimal failure info so the test still appears in report
                import re
                match = re.search(r'C\d+', test_name)
                test_case_id = match.group(0) if match else 'N/A'
                
                failures.append({
                    'test_case_id': test_case_id,
                    'test_case_uuid': test.get('testCaseUuid'),
                    'test_name': test_name,
                    'status': test.get('status', ''),
                    'screenshot_number': 'N/A',
                    'failed_command': 'N/A',
                    'error_message': f'Error fetching details: {str(e)}',
                    'screenshot_url': None,
                    'screenshot_urls': []
                })
    
    print(f"Total failures found and processed: {len(failures)}")
    return failures

def process_failures_minimal(test_cases):
    """Optimized version for Rerun Failed Tests - no execution details, screenshots, or commands"""
    failures = []
    
    print(f"\n=== PROCESSING FAILURES (MINIMAL - OPTIMIZED) ===")
    print(f"Total test cases: {len(test_cases)}")
    
    for test in test_cases:
        # Check if test failed
        status = test.get('status', '')
        if status in ['Failed', 'failed', 'error', 'Error']:
            # Extract Test Case ID from name (C followed by numbers)
            test_name = test.get('name', 'Unknown Test')
            import re
            match = re.search(r'C\d+', test_name)
            test_case_id = match.group(0) if match else 'N/A'
            
            # Get basic info only - no API calls for execution details
            test_case_uuid = test.get('testCaseUuid')
            
            # Get error message from test object if available
            error_message = test.get('errorMessage', test.get('message', test.get('failureMessage', 'Test failed')))
            
            failure_info = {
                'test_case_id': test_case_id,
                'test_case_uuid': test_case_uuid,
                'test_name': test_name,
                'status': status,
                'screenshot_number': 'N/A',
                'failed_command': 'N/A',
                'error_message': error_message if error_message else 'Test failed',
                'screenshot_url': None,
                'screenshot_urls': []
            }
            
            failures.append(failure_info)
            print(f"Found failure: {test_case_id} - {test_name}")
    
    print(f"Total failures found: {len(failures)} (processed in minimal mode)")
    return failures

def fetch_test_details(run_id, test_id, api_key, base_url):
    """Fetch detailed information for a specific test"""
    try:
        headers = {
            'auth-token': api_key
        }
        
        url = f'{base_url}/runs/{run_id}/testcases/{test_id}'
        response = requests.get(url, headers=headers)
        
        if response.status_code == 200:
            return response.json()
        else:
            print(f"Error fetching test details: {response.status_code} - {response.text}")
            return None
    
    except Exception as e:
        print(f"Exception in fetch_test_details: {str(e)}")
        return None

def fetch_step_screenshot(run_id, test_case_uuid, execution_uuid, step_number, api_key, base_url):
    """Fetch screenshot URL for a specific step"""
    try:
        headers = {
            'auth-token': api_key
        }
        
        # Try to get screenshot from step details endpoint
        url = f'{base_url}/runs/{run_id}/testcases/{test_case_uuid}/executions/{execution_uuid}/steps/{step_number}'
        print(f"\n=== FETCHING SCREENSHOT FOR STEP {step_number} ===")
        print(f"URL: {url}")
        response = requests.get(url, headers=headers)
        print(f"Response Status: {response.status_code}")
        
        if response.status_code == 200:
            step_data = response.json()
            print(f"Step data keys: {list(step_data.keys())}")
            data = step_data.get('data', step_data)
            print(f"Data keys: {list(data.keys()) if isinstance(data, dict) else 'Not a dict'}")
            print(f"Full step data: {json.dumps(step_data, indent=2, default=str)[:1000]}")
            
            # Check for screenshot in response - check all possible fields
            for field in ['screenshot', 'screenshotUrl', 'screenshotLink', 'image', 'imageUrl', 'screenshotPath', 'screenshotS3Url']:
                if data.get(field):
                    screenshot_url = data.get(field)
                    print(f"Found screenshot URL in field '{field}': {screenshot_url}")
                    return screenshot_url
            
            print("No screenshot URL found in any expected field")
        else:
            print(f"Step details endpoint returned: {response.status_code}")
            print(f"Response text: {response.text[:500]}")
        
        return None
    except Exception as e:
        print(f"Exception fetching step screenshot: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

# Create a session for connection pooling and reuse
_screenshot_session = requests.Session()

def validate_screenshot_url(url):
    """Validate screenshot URL with fallback to highlighted version
    
    Returns:
        tuple: (validated_url or None, is_available)
    """
    if not url:
        return None, False
    
    try:
        # First attempt: Try the original screenshot URL (reduced timeout for speed)
        response = _screenshot_session.head(url, timeout=2)
        if response.status_code == 200:
            return url, True
        
        # Fallback: Try the highlighted version
        if url.endswith('.png'):
            highlighted_url = url.replace('.png', '_highlighted.png')
            response = _screenshot_session.head(highlighted_url, timeout=2)
            if response.status_code == 200:
                return highlighted_url, True
        
        return None, False
    
    except Exception as e:
        return None, False

def validate_screenshots_batch(urls):
    """Validate multiple screenshot URLs concurrently for better performance
    
    Args:
        urls: List of screenshot URLs to validate
        
    Returns:
        list: List of validated URLs (or None for unavailable screenshots)
    """
    if not urls:
        return []
    
    print(f"Validating {len(urls)} screenshots concurrently...")
    validated_urls = []
    
    # Use ThreadPoolExecutor for concurrent requests
    with ThreadPoolExecutor(max_workers=10) as executor:
        # Submit all validation tasks
        future_to_url = {executor.submit(validate_screenshot_url, url): url for url in urls}
        
        # Collect results in order
        url_results = {}
        for future in as_completed(future_to_url):
            original_url = future_to_url[future]
            try:
                validated_url, is_available = future.result()
                url_results[original_url] = validated_url if is_available else None
            except Exception as e:
                print(f"Error validating {original_url}: {str(e)}")
                url_results[original_url] = None
        
        # Return results in original order
        validated_urls = [url_results.get(url) for url in urls]
    
    available_count = sum(1 for url in validated_urls if url is not None)
    print(f"Screenshot validation complete: {available_count}/{len(urls)} available")
    return validated_urls

def fetch_execution_details(run_id, test_case_uuid, execution_uuid, api_key, base_url):
    """Fetch detailed execution information including steps"""
    try:
        headers = {
            'auth-token': api_key
        }
        
        url = f'{base_url}/runs/{run_id}/testcases/{test_case_uuid}/executions/{execution_uuid}'
        response = requests.get(url, headers=headers)
        
        if response.status_code == 200:
            response_json = response.json()
            
            # Handle different response structures
            if isinstance(response_json, dict):
                # Check if data is present and is a list (steps array)
                data = response_json.get('data', response_json)
                if isinstance(data, list):
                    return data  # Return the list of steps directly
                else:
                    return data  # Return dict as is
            elif isinstance(response_json, list):
                return response_json  # Return list of steps directly
            else:
                return []
        else:
            print(f"Error fetching execution details: {response.status_code}")
            return None
    
    except Exception as e:
        print(f"Exception in fetch_execution_details: {str(e)}")
        return None

if __name__ == '__main__':
    # Create reports directory if it doesn't exist
    os.makedirs('reports', exist_ok=True)
    
    app.run(debug=True, host='0.0.0.0', port=5000)
