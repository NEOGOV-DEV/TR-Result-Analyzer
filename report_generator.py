import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image as XLImage
from openpyxl.comments import Comment
from datetime import datetime
import requests
import os
from io import BytesIO

class ReportGenerator:
    def __init__(self):
        self.reports_dir = 'reports'
        os.makedirs(self.reports_dir, exist_ok=True)
    
    def generate_excel_report(self, run_id, run_details, failures, api_key=None, suite_name=None):
        """Generate an Excel report with failure details"""
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Test Failure Report"
        
        # Define styles
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=12)
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Add report header
        ws.merge_cells('A1:F1')
        ws['A1'] = "Test Rigor Failure Report"
        ws['A1'].font = Font(bold=True, size=16)
        ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
        
        # Add run information
        ws['A3'] = "Suite Name:"
        ws['B3'] = suite_name if suite_name else 'N/A'
        ws['A3'].font = Font(bold=True)
        
        ws['A4'] = "Run ID:"
        ws['B4'] = run_id
        ws['A4'].font = Font(bold=True)
        
        ws['A5'] = "Generated On:"
        ws['B5'] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        ws['A5'].font = Font(bold=True)
        
        ws['A6'] = "Total Failed Tests:"
        ws['B6'] = len(failures)
        ws['A6'].font = Font(bold=True)
        ws['B6'].font = Font(color="FF0000", bold=True)
        
        if run_details:
            ws['D4'] = "Run Name:"
            # Get the custom name from run details
            run_name = run_details.get('customName', 'N/A')
            ws['E4'] = run_name
            ws['D4'].font = Font(bold=True)
            
            ws['D5'] = "Status:"
            ws['E5'] = run_details.get('status', 'N/A')
            ws['D5'].font = Font(bold=True)
        
        # Add table headers
        headers = ['Test Case ID', 'Test Name', 'Status', 'Failed Screenshot#', 'Failed Command', 'Error Message', 'Screenshot']
        header_row = 8
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_num)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add failure data
        current_row = header_row + 1
        
        for failure in failures:
            screenshot_urls = failure.get('screenshot_urls', [])
            num_screenshots = len(screenshot_urls) if screenshot_urls else 1
            start_row = current_row
            
            # Add each screenshot in a separate row
            for i in range(num_screenshots):
                if i == 0:
                    # First row - add all data
                    ws.cell(row=current_row, column=1).value = failure.get('test_case_id', 'N/A')
                    ws.cell(row=current_row, column=2).value = failure.get('test_name', 'N/A')
                    ws.cell(row=current_row, column=3).value = failure.get('status', 'N/A')
                    ws.cell(row=current_row, column=4).value = failure.get('screenshot_number', 'N/A')
                    failed_command = failure.get('failed_command', '').strip()
                    if failed_command:
                        # Process multi-line commands and add "No command found" for empty steps
                        lines = failed_command.split('\n')
                        processed_lines = []
                        for line in lines:
                            line = line.strip()
                            if line:
                                # Check if line is just "Step X:" with no command
                                if line.endswith(':') and 'Step' in line:
                                    processed_lines.append(f"{line} No command found")
                                else:
                                    processed_lines.append(line)
                        failed_command = '\n'.join(processed_lines)
                    else:
                        failed_command = 'No command found'
                    ws.cell(row=current_row, column=5).value = failed_command
                    ws.cell(row=current_row, column=6).value = failure.get('error_message', 'N/A')
                
                # Add screenshot link for this row
                if screenshot_urls and i < len(screenshot_urls):
                    cell = ws.cell(row=current_row, column=7)
                    cell.value = f"Screenshot {i + 1}"
                    cell.hyperlink = screenshot_urls[i]
                    cell.font = Font(color="0000FF", underline="single")
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                else:
                    cell = ws.cell(row=current_row, column=7)
                    cell.value = 'No screenshot available'
                    cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Apply borders to all cells in the row
                for col in range(1, 8):
                    ws.cell(row=current_row, column=col).border = border
                    ws.cell(row=current_row, column=col).alignment = Alignment(vertical='top', wrap_text=True)
                
                current_row += 1
            
            # Merge cells for columns 1-6 if there are multiple screenshots
            if num_screenshots > 1:
                for col in range(1, 7):
                    ws.merge_cells(start_row=start_row, start_column=col, end_row=current_row - 1, end_column=col)
                    # Center the content vertically in merged cells
                    ws.cell(row=start_row, column=col).alignment = Alignment(vertical='center', wrap_text=True)
        
        # Adjust column widths
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 12
        ws.column_dimensions['D'].width = 18
        ws.column_dimensions['E'].width = 40
        ws.column_dimensions['F'].width = 50
        ws.column_dimensions['G'].width = 30  # Wider to accommodate images
        
        # Save the workbook
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"TestRigor_FailureReport_{run_id}_{timestamp}.xlsx"
        filepath = os.path.join(self.reports_dir, filename)
        
        wb.save(filepath)
        
        return filename
    
    def download_screenshot(self, url, test_id):
        """Download screenshot from URL"""
        try:
            response = requests.get(url)
            if response.status_code == 200:
                filename = f"screenshot_{test_id}.png"
                filepath = os.path.join(self.reports_dir, filename)
                with open(filepath, 'wb') as f:
                    f.write(response.content)
                return filepath
            return None
        except Exception as e:
            print(f"Error downloading screenshot: {str(e)}")
            return None
